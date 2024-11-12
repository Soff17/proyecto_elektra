import fitz
import re
import os
from funciones import watson_discovery as wd
from funciones import image_storage as st
from funciones import elastict_search as es
import io
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor
from openpyxl.styles import PatternFill
from unidecode import unidecode
from datetime import datetime
from PIL import Image

sku_pattern = re.compile(r'Sku:\s*(\S+)')
sku_pattern_2 = re.compile(r'Sku de referencia:\s*(\S+)')
sku_pattern_3 = re.compile(r'Sku[´\'s]{0,4} de referencia:\s*(\S+)|Sku:\s*(\S+)')
vigencia_pattern = re.compile(r'Vigencia:\s*(.+)')
delete_pattern = re.compile(r'(Da clic aquí)s?')
delete_pattern2 = re.compile(r'(¡Cómpralo ya!)s?')
precio_pattern = re.compile(r'Contado\s*(\S+)')
precio_pattern2 = re.compile(r'^\$\d{1,3}(,\d{3})+(\.\d{2})?$', re.MULTILINE)
bono_pattern = re.compile(r'(¡ B O N O  D E  R E G A L O)s?')
bono_pattern2 = re.compile(r'D E\s*(.+)')

titulos = []
subtitulos = []
info = []
urls = []
skus = []
vigencias = []
precios = []
#cupones = []
nueva_data_productos = []

# Definir listas para almacenar la información que vamos a exportar
data_reporte = {
    'SKU': [],
    'Subtítulo': [],
    'Descripción': [],
    'Vigencia': [],
    'URL': []
}

def elimanr_ultimas_paginas(pdf, output_path):
    doc = fitz.open(pdf)

    number_of_pages = doc.page_count

    page_to_delete = [number_of_pages - 1, number_of_pages - 2, number_of_pages - 3]

    page_to_delete.sort(reverse=True)

    for page in page_to_delete:
        doc.delete_page(page)

    # El output path tiene que ser el nombre del archvio
    # Ejemplo: ./pdf_buffer.pdf
    doc.save(output_path)
    doc.close

def nombre_de_categoria(font_size, font_flags):
    if (font_size > 43) and font_flags == 20:
        return True
    return False

def nombre_del_producto(font_size, font_flags, categoria):
    if categoria == "Planes":
        return (font_size > 28 and font_size < 34) and (font_flags == 20 or font_flags == 4)
    else:
        return (font_size > 31.5 and font_size < 41.0) and (font_flags == 20 or font_flags == 4)


def extraer_informacion(page):
    blocks = page.get_text("dict", sort=True)["blocks"]

    inicio_productos = False
    inicio_producto = False
    fin_producto = False
    datos = ''
    sku_positions = []

    # Nuevo patrón para capturar precios precedidos por "Contado"
    precio_pattern3 = re.compile(r'Contado\s*\$\d{1,3}(,\d{3})*(\.\d{2})?')

    # Determinar la categoría antes de procesar el resto de la información
    categoria = extraer_categoria(page)
    nombre_producto_actual = ""  # Variable para capturar temporalmente el nombre del producto

    for block in blocks:
        if 'lines' in block:
            for line in block['lines']:
                for span in line['spans']:
                    text = span['text'].strip()
                    text_size = span['size']
                    text_flags = span['flags']
                    text_y_position = span['bbox'][1]

                    # Detección de categoría
                    if nombre_de_categoria(text_size, text_flags) and not inicio_producto:
                        if not inicio_productos and len(titulos) > 0:
                            titulos[-1] += " " + text
                        else:
                            titulos.append(text)

                    # Acumulación del nombre de producto en múltiples líneas
                    elif nombre_del_producto(text_size, text_flags, categoria):
                        if inicio_producto:
                            nombre_producto_actual += " " + text
                        else:
                            nombre_producto_actual = text
                            inicio_producto = True
                        inicio_productos = True

                    # Detección de SKU y consolidación del nombre de producto
                    elif sku_pattern.findall(text):
                        sku = sku_pattern.findall(text)[0].replace(".", "")
                        skus.append(sku)
                        sku_positions.append((sku, text_y_position))
                        fin_producto = True

                        # Asignar el nombre del producto acumulado a subtítulos si está disponible
                        if nombre_producto_actual:
                            subtitulos.append(nombre_producto_actual.strip())  # Limpiar cualquier espacio adicional
                            nombre_producto_actual = ""  # Limpiar para el próximo producto
                        else:
                            subtitulos.append("Producto")
                            precios.append("Pago de contado: NA")

                    elif sku_pattern_2.findall(text):
                        sku = sku_pattern_2.findall(text)[0].replace(".", "")
                        skus.append(sku)
                        sku_positions.append((sku, text_y_position))
                        fin_producto = True
                        inicio_producto = True
                        
                        # Consolidación para este caso de SKU alternativo
                        if nombre_producto_actual:
                            subtitulos.append(nombre_producto_actual.strip())
                            nombre_producto_actual = ""
                        else:
                            subtitulos.append("Producto")
                            precios.append("Pago de contado: NA")

                    elif sku_pattern_3.findall(text):
                        text = text.replace('Sku´s de referencia: ', '')
                        skus_encontrados = re.findall(r'\d+', text)

                        if skus_encontrados:
                            primer_sku = skus_encontrados[0]
                            if primer_sku in skus and len(skus_encontrados) > 1:
                                segundo_sku = skus_encontrados[1]
                                sku_combination = f"{primer_sku} y {segundo_sku}"
                                skus.append(sku_combination)
                                sku_positions.append((sku_combination, text_y_position))
                            elif primer_sku not in skus:
                                skus.append(primer_sku)
                                sku_positions.append((primer_sku, text_y_position))
                            elif len(skus_encontrados) > 1:
                                segundo_sku = skus_encontrados[1]
                                skus.append(segundo_sku)
                                sku_positions.append((segundo_sku, text_y_position))

                        fin_producto = True
                        inicio_producto = True
                        
                        # Consolidación del nombre del producto acumulado
                        if nombre_producto_actual:
                            subtitulos.append(nombre_producto_actual.strip())
                            nombre_producto_actual = ""
                        else:
                            subtitulos.append("Producto")
                            precios.append("Pago de contado: NA")

                    # Captura de vigencia
                    elif vigencia_pattern.findall(text) and fin_producto and inicio_producto:
                        vigencias.append(text)
                        info.append(datos)
                        datos = ''
                        fin_producto = False
                        inicio_producto = False

                    # Eliminar información extra
                    elif delete_pattern.findall(text) or delete_pattern2.findall(text):
                        continue
                    
                    # Captura de precios
                    elif precio_pattern3.findall(text):
                        precio = precio_pattern3.findall(text)[0]
                        precios.append(f"Pago de contado: {precio}")

                    # Captura de información del producto
                    else:
                        if datos == '':
                            datos = text
                        else:
                            datos += " " + text

    return sku_positions  # Devolver las posiciones de los SKUs para asignar imágenes



def convertir_a_jpeg(imagen_bytes):
    with Image.open(io.BytesIO(imagen_bytes)) as img:
        img_rgb = img.convert('RGB')  # Convertir a RGB si es necesario
        buffer = io.BytesIO()
        img_rgb.save(buffer, format="JPEG")  # Guardar como JPEG
        buffer.seek(0)
        return buffer.getvalue()  # Devolver bytes de la imagen en JPEG
    
def extraer_imagenes_orden(bucket_name, bucket_folder, page, doc, sku_positions):
    images = page.get_image_info(hashes=True, xrefs=True)
    imagenes = []

    for img in images:
        xref = img['xref']
        if xref > 0:
            if img['width'] > 269.5 and img['height'] > 269.5:
                bbox_img = img['bbox']
                imagenes.append((xref, bbox_img))

    # Ordenamos las imágenes por su posición Y en la página
    images_sorted = sorted(imagenes, key=lambda img: img[1][3], reverse=False)
    
    # Crear lista para las subidas
    uploads = []
    count = 0

    for xref, bbox in images_sorted:
        base_image = doc.extract_image(xref)
        image_bytes = base_image["image"]
        ext = base_image["ext"]

        if ext != "jpeg":
            # Convertir a JPEG si la extensión no es "jpeg"
            image_bytes = convertir_a_jpeg(image_bytes)

        # Encontrar el SKU más cercano basado en la posición Y
        closest_sku = find_closest_sku(sku_positions, bbox[3])

        if closest_sku:
            image_name = f"{closest_sku}.jpeg"
        else:
            image_name = f"producto_{count+1}.{ext}"

        # Obtener la carpeta de descargas del usuario
        downloads_folder = get_downloads_folder()

        # Crear el directorio 'imagenes' dentro de la carpeta Descargas si no existe
        ruta_imagenes = os.path.join(downloads_folder, 'imagenes')
        if not os.path.exists(ruta_imagenes):
            os.makedirs(ruta_imagenes)

        # Guardar la imagen en la nueva carpeta de Descargas
        ruta_imagen = os.path.join(ruta_imagenes, image_name)
        with open(ruta_imagen, "wb") as f:
            f.write(image_bytes)
        
        # Preparar la subida con los datos de la imagen
        image_buffer = io.BytesIO(image_bytes)
        uploads.append((bucket_name, bucket_folder, image_name, image_buffer))

        print(f"Preparado para subir imagen {image_name} al bucket.")
        count += 1

    # Usar hilos para subir las imágenes
    with ThreadPoolExecutor() as executor:
        futures = [
            #executor.submit(st.upload_image_buffer, bucket_name, bucket_folder, image_name, image_buffer)
            executor.submit(bucket_name, bucket_folder, image_name, image_buffer)
            for bucket_name, bucket_folder, image_name, image_buffer in uploads
        ]

        # Capturar errores en las subidas
        for future in futures:
            try:
                future.result()
            except Exception as e:
                print(f"Error al subir la imagen: {e}")

    # Si no se encontraron imágenes para algún SKU, usa 'default.jpeg' desde Descargas/imagenes
    for sku, _ in sku_positions:
        ruta_imagen_sku = os.path.join(ruta_imagenes, f"{sku}.jpeg")
        if not os.path.exists(ruta_imagen_sku):
            ruta_default = './default.jpeg'
            if os.path.exists(ruta_default):
                ruta_imagen = os.path.join(ruta_imagenes, f"{sku}.jpeg")
                
                # Copiar localmente la imagen predeterminada
                with open(ruta_default, "rb") as f_default, open(ruta_imagen, "wb") as f_sku:
                    default_image_bytes = f_default.read()
                    f_sku.write(default_image_bytes)
                print(f"Imagen predeterminada asignada al SKU {sku} localmente.")

                # Subir la imagen predeterminada al bucket usando hilos
                default_image_buffer = io.BytesIO(default_image_bytes)
                with ThreadPoolExecutor() as executor:
                    #executor.submit(st.upload_image_buffer, bucket_name, bucket_folder, f"{sku}.jpeg", default_image_buffer)
                    executor.submit(bucket_name, bucket_folder, f"{sku}.jpeg", default_image_buffer)
                print(f"Imagen predeterminada asignada y subida para el SKU {sku}.")

            else:
                print(f"Imagen predeterminada no encontrada.")

def get_default_image_buffer_from_local(default_image_path="./default.jpeg"):
    try:
        # Abrir la imagen en modo binario
        with open(default_image_path, "rb") as default_image_file:
            # Cargar la imagen en un buffer
            image_buffer = io.BytesIO(default_image_file.read())
            image_buffer.seek(0)  # Reiniciar el puntero del buffer al inicio
            print(f"Imagen predeterminada '{default_image_path}' cargada exitosamente desde la raíz del proyecto.")
            return image_buffer
    except FileNotFoundError:
        print(f"Imagen predeterminada '{default_image_path}' no encontrada en la raíz del proyecto.")
        return None

def find_closest_sku(sku_positions, image_y_position):
    closest_sku = None
    min_distance = float('inf')
    # Set a threshold to avoid matching distant images (e.g., set to 200)
    threshold = 200  

    for sku, sku_position in sku_positions:
        distance = abs(sku_position - image_y_position)
        if distance < min_distance and distance < threshold:
            closest_sku = sku
            min_distance = distance
            sku_positions = [tupla for tupla in sku_positions if tupla[0] != closest_sku]

    return closest_sku

def get_urls(page):
    links = page.get_links()
    urls_with_rect = []

    # Extract URL and Rect from each link and store in a list of tuples
    for link in links:
        if 'uri' in link and 'from' in link:
            url = link['uri']
            rect = link['from']
            coordinates = [rect[0], rect[1], rect[2], rect[3]]
            urls_with_rect.append((url, coordinates))

    # Sort the URLs based on rectangular coordinates (x, y)
    urls_sorted = sorted(urls_with_rect, key=lambda x: (x[1][1], x[1][0]))

    # Extract only the sorted URLs
    urls_sorted = [url for url, _ in urls_sorted]

    # Iterate through each SKU and match with URLs
    for i in range(len(skus)):
        sku = skus[i]
        original_url = urls_sorted[i] if i < len(urls_sorted) else None
        matched_url = None

        for url in urls_sorted:
            if sku in url:
                matched_url = url
                break

        # If no matched URL found, use the original URL instead of "dummy-url"
        if not matched_url:
            matched_url = original_url if original_url else "dummy-url"

        urls.append(matched_url)

def guardar_informacion_a_elasticsearch(name_file, data, bucket_name, carpeta_documentos_correcciones_bucket, carpeta_documentos_elastic_bucket):   
    # Generar el contenido del archivo como una cadena
    contenido_txt = "\n".join(data)

    # Validar si la URL contiene el SKU y si la imagen existe
    sku = data[0].replace("Sku: ", "")
    url = data[-1].replace("Url: ", "")
    imagen_existe = tiene_imagen(sku)
    subtitulo = data[2]
    categoria = data[1].replace("Categoria:", "").strip().lower()  # Remover "Categoria:" y espacios en blanco

    # Definir la estructura requerida
    estructura_requerida = ["Sku:", "Categoria:", "Producto:", "Pago semanal:", "Descuento:", "Pago de contado:", "Vigencia:"]

    # Expresión regular para validar el formato de "Pago semanal"
    # Expresión regular para validar el formato de "Pago semanal"
    pago_semanal_pattern = re.compile(
        r'Pago semanal: \$?\d+ x \d+ semanas \$?\d+(,\d{3})* de pago inicial|Pago semanal: \d+ \$ x \d+ semanas con \d+% de enganche'
    )


    # Verificar si el producto cumple con la estructura requerida
    def cumple_estructura(data):
        for campo in estructura_requerida:
            if not any(campo in item for item in data):
                return False
        return True

    # Validaciones para enviar a correcciones
    if (
        not imagen_existe or 
        subtitulo.startswith("Producto: Promoción") or 
        (not cumple_estructura(data) and "equipos" not in categoria and "planes" not in categoria) or 
        not categoria or  # Validar si la categoría está vacía después de quitar espacios
        (categoria != "planes" and not any(pago_semanal_pattern.match(line) for line in data))  # Validar estructura de "Pago semanal" si no es "planes"
    ):
        guardar_en_correcciones(name_file, contenido_txt, bucket_name, carpeta_documentos_correcciones_bucket)
        print(f"'{name_file}.txt' guardado en 'correcciones' debido a validación fallida.")
        return


    # Crear una lista de tareas para ejecutar en paralelo
    tasks = []

    # Subida a Elasticsearch
    documento = {
        "titulo": f"{name_file}",
        "contenido": contenido_txt
    }
    tasks.append(("elasticsearch", documento, name_file))

    # Guardar el archivo localmente
    downloads_folder = get_downloads_folder()
    ruta_output_files = os.path.join(downloads_folder, 'documentos_elastic')
    
    if not os.path.exists(ruta_output_files):
        os.makedirs(ruta_output_files)

    file_path = os.path.join(ruta_output_files, f"{name_file}.txt")
    with open(file_path, 'w', encoding='utf-8') as file:
        file.write(contenido_txt)

    # Subida al bucket en la carpeta documentos_elastic
    buffer = io.BytesIO(contenido_txt.encode('utf-8'))
    buffer.seek(0)  # Asegurarse de que el buffer esté al inicio
    tasks.append(("storage", bucket_name, carpeta_documentos_elastic_bucket, f"{name_file}.txt", buffer))

    # Ejecutar las subidas en paralelo
    with ThreadPoolExecutor() as executor:
        futures = []
        
        for task in tasks:
            if task[0] == "elasticsearch":
                documento, name_file = task[1], task[2]
                futures.append(executor.submit("catalogo", name_file, documento))
            elif task[0] == "storage":
                _, bucket, folder, filename, buffer = task
                futures.append(executor.submit(bucket, folder, filename, buffer))

        # Verificar y manejar posibles errores en las subidas
        for future in futures:
            try:
                future.result()
            except Exception as e:
                print(f"Error al subir documento: {e}")

    print(f'Se subió "{name_file}.txt" a Elasticsearch y almacenamiento exitosamente.')



def tiene_imagen(sku):
    # Verifica si existe una imagen para el SKU
    extensiones_imagen = ['jpeg']
    downloads_folder = get_downloads_folder()
    ruta_imagenes = os.path.join(downloads_folder, 'imagenes')
    for ext in extensiones_imagen:
        if os.path.exists(f"{ruta_imagenes}/{sku}.{ext}"):
            return True
    return False

from concurrent.futures import ThreadPoolExecutor
import io
import os

def guardar_en_correcciones(name_file, contenido, bucket_name, carpeta_documentos_correcciones_bucket):
    # Obtener la carpeta de descargas del usuario
    downloads_folder = get_downloads_folder()
    ruta_correcciones = os.path.join(downloads_folder, 'correcciones')
    
    if not os.path.exists(ruta_correcciones):
        os.makedirs(ruta_correcciones)

    # Ruta completa para guardar el archivo localmente
    file_path = os.path.join(ruta_correcciones, f"{name_file}.txt")
    
    # Crear un buffer para la subida a storage
    buffer = io.BytesIO(contenido.encode('utf-8'))
    buffer.seek(0)  # Asegurar que el buffer esté al inicio

    # Definir las tareas de guardado local y subida en paralelo
    def guardar_local():
        with open(file_path, 'w', encoding='utf-8') as file:
            file.write(contenido)
        print(f"'{name_file}.txt' guardado en 'correcciones' para revisión.")

    def subir_a_storage():
        #st.upload_text_buffer(bucket_name, carpeta_documentos_correcciones_bucket, f"{name_file}.txt", buffer)
        print(f"'{name_file}.txt' también subido a storage en 'correcciones'.")

    # Ejecutar las tareas en paralelo usando ThreadPoolExecutor
    with ThreadPoolExecutor() as executor:
        futures = [executor.submit(guardar_local), executor.submit(subir_a_storage)]

        # Manejar posibles errores en ambas tareas
        for future in futures:
            try:
                future.result()
            except Exception as e:
                print(f"Error al ejecutar tarea: {e}")

def guardar_informacion_a_discovery(titulo, name_file, data):
    # Reemplazar espacios con guiones bajos
    titulo = re.sub(r'[^A-Za-zÁÉÍÓÚáéíóúÑñ\s]', '', titulo).strip()
    titulo = unidecode(titulo).replace(" ", "_")  # Quitar acentos con unidecode
    
    # Generar el contenido del archivo como una cadena
    contenido_txt = "\n".join(data)
    
    # Subir directamente a IBM Watson Discovery
    from funciones import watson_discovery as wd  # Importar Watson Discovery
    
    # Usar una función similar a `añadir_documento` para subir el contenido
    # wd.añadir_documento_desde_contenido(contenido_txt, f"{titulo} {name_file}.txt", 'text/plain')
    print(f'se está subiendo "{titulo} {name_file}.txt"')

    # Obtener la carpeta de descargas del usuario
    downloads_folder = get_downloads_folder()

    # Crear el directorio 'output_files' dentro de la carpeta Descargas si no existe
    ruta_output_files = os.path.join(downloads_folder, 'output_files')
    if not os.path.exists(ruta_output_files):
        os.makedirs(ruta_output_files)

    # Definir la ruta completa del archivo
    file_path = os.path.join(ruta_output_files, f"{titulo} {name_file}.txt")

    # Guardar el contenido en un archivo local
    with open(file_path, 'w', encoding='utf-8') as file:
        file.write(contenido_txt)
        
def particion_pdf(pdf_buffer, bucket_name, bucket_folder):
    # Abre el PDF desde el buffer en memoria
    doc = fitz.open(stream=pdf_buffer, filetype="pdf")
    buffer_categoria = ""
    
    # Obtener la carpeta de descargas del usuario
    downloads_folder = get_downloads_folder()

    # Crear la carpeta 'pdfs_finales' dentro de la carpeta Descargas si no existe
    ruta_pdfs_finales = os.path.join(downloads_folder, 'pdfs_finales')
    if not os.path.exists(ruta_pdfs_finales):
        os.makedirs(ruta_pdfs_finales)

    for num_page in range(doc.page_count):
        # Crear un nuevo PDF con solo una página
        doc_pagina = fitz.open()
        doc_pagina.insert_pdf(doc, from_page=num_page, to_page=num_page)
        
        # Extraer el nombre de la categoría de la página
        page = doc.load_page(num_page)
        categoria = extraer_categoria(page)

        # Sanitizar el nombre de la categoría para ser usado en el nombre de archivo
        nombre_categoria_sanitizado = sanitizar_nombre_categoria(categoria)

        # Nombre del archivo para la página basado en la categoría
        nombre_archivo_pdf = f"{nombre_categoria_sanitizado}.pdf"
        
        # Guardar localmente en la carpeta 'pdfs_finales'
        ruta_local_pdf = os.path.join(ruta_pdfs_finales, nombre_archivo_pdf)

        print(ruta_local_pdf)

        # Guardar el PDF de una página en la ruta local
        doc_pagina.save(ruta_local_pdf)
        print(f"Página {num_page + 1} guardada como {ruta_local_pdf}")
        
        # Crear un buffer de bytes para almacenar el PDF en memoria
        pdf_buffer_output = io.BytesIO()
        doc_pagina.save(pdf_buffer_output)
        pdf_buffer_output.seek(0)  # Regresar al inicio del buffer

        # Subir el PDF al bucket directamente desde el buffer
        #st.upload_pdf_buffer(bucket_name, bucket_folder, nombre_archivo_pdf, pdf_buffer_output)
        print(f"PDF {nombre_archivo_pdf} subido exitosamente al bucket.")

        doc_pagina.close()
        if nombre_categoria_sanitizado == "nombre_temporal":
            join_pdfs(os.path.join(ruta_pdfs_finales, f"{buffer_categoria}.pdf"), ruta_local_pdf)
        buffer_categoria = nombre_categoria_sanitizado

    doc.close()


# Lista global para almacenar las categorías extraídas
categorias_extraidas = []

def extraer_categoria(page):
    # Lógica para extraer el nombre de la categoría desde la página
    blocks = page.get_text("dict", sort=True)["blocks"]
    
    for block in blocks:
        if 'lines' in block:
            for line in block['lines']:
                for span in line['spans']:
                    text = span['text'].strip()
                    text_size = span['size']
                    text_flags = span['flags']

                    # Verificar si el texto coincide con el nombre de la categoría
                    if nombre_de_categoria(text_size, text_flags):
                        # Añadir la categoría detectada a la lista
                        categorias_extraidas.append(text)
                        return text
    return "Categoria_Desconocida"


def join_pdfs(first_pdf_path, second_pdf_path):
    doc1 = fitz.open(first_pdf_path)
    doc2 = fitz.open(second_pdf_path)

    # Insertar el contenido del segundo documento en el primero
    doc1.insert_pdf(doc2)

    # Guardar los cambios en un archivo temporal primero
    temp_path = "temp_file.pdf"
    doc1.save(temp_path, incremental=False)  # Guardar sin modo incremental
    doc1.close()
    doc2.close()
    os.remove(second_pdf_path)

    # Mover el archivo temporal al archivo original, sobrescribiéndolo
    os.replace(temp_path, first_pdf_path)
    print(f"El PDF combinado ha sido guardado sobre: {first_pdf_path}")

def sanitizar_nombre_categoria(categoria):
    # Reemplazar solo los caracteres no válidos, mantener los espacios y convertirlos a guiones bajos
    categoria_sanitizada = re.sub(r'[^A-Za-zÁÉÍÓÚáéíóúÑñ\s]', '', categoria).strip()

    nombre_con_guiones_bajos = unidecode(categoria_sanitizada).replace(" ", "_")
    # Si hay varios guiones bajos, cortar en el primer guión bajo
    nombre_final = nombre_con_guiones_bajos.split('_')[0]
    
    # Excepciones para nombres específicos
    if nombre_final == "Home":
        nombre_final = "Home_Audio"
    elif nombre_final == "Linea":
        nombre_final = "Linea_Blanca"
    elif not nombre_final:
        return "nombre_temporal"
    
    return nombre_final

def ajustar_longitudes_listas():
    # Encontrar la longitud máxima entre todas las listas
    max_length = max(len(skus), len(subtitulos), len(info), len(vigencias), len(urls), len(nueva_data_productos))

    # Rellenar con valores vacíos en caso de que alguna lista sea más corta
    skus.extend([""] * (max_length - len(skus)))
    subtitulos.extend([""] * (max_length - len(subtitulos)))
    info.extend([""] * (max_length - len(info)))
    vigencias.extend([""] * (max_length - len(vigencias)))
    urls.extend([""] * (max_length - len(urls)))
    
    # Ajustar también la lista 'nueva_data_productos' para que coincida
    nueva_data_productos.extend([["", "", "", "", "", "", ""]] * (max_length - len(nueva_data_productos)))

ultima_categoria = "Categoria: null" 

def procesar_pdf(pdf_buffer, bucket_name, carpeta_imagenes_bucket, carpeta_pdfs_bucket, carpeta_documentos_correcciones_bucket, carpeta_documentos_elastic_bucket, carpeta_reportes_bucket ):
    nueva_data_productos.clear()
    # Abre el PDF desde el buffer en memoria
    doc = fitz.open(stream=pdf_buffer, filetype="pdf")

    for page_num in range(doc.page_count):
        page = doc.load_page(page_num)

        titulos.clear()
        subtitulos.clear()
        info.clear()
        urls.clear()
        skus.clear()
        vigencias.clear()
        #cupones.clear()  # También limpia la lista de cupones si es necesario

        # Llamada a extraer_informacion para obtener las posiciones de los SKUs
        sku_positions = extraer_informacion(page)

        if len(titulos) == 0 or len(info) == 0:
            print(f"No se encontró información en la página {page_num + 1}. Continuando con la siguiente página.")
            continue

        get_urls(page)

        # Pasar sku_positions a la función extraer_imagenes_orden
        extraer_imagenes_orden(bucket_name, carpeta_imagenes_bucket, page, doc, sku_positions)

        for i in range(max(len(subtitulos), len(info), len(skus), len(vigencias), len(urls), len(precios))):
            sku = skus[i] if i < len(skus) else ""
            vigencia = vigencias[i] if i < len(vigencias) else ""
            subtitulo = subtitulos[i] if i < len(subtitulos) else ""
            datos_producto = info[i] if i < len(info) else ""
            subtitulo = subtitulo.replace('con hasta', '')

            #if titulos[0] == "50%":
                #titulos[0] = "Línea Blanca"

            if titulos[0] == "Planes":
                subtitulo = re.sub(r'^.*?Llévate un', '', subtitulo).strip()
                subtitulo = re.sub(r'con\s.*', '', subtitulo).strip()

                #cupon = "Sin Cupones"

                # Corregir el formato de "95 $ Desde x 128 semanas" a "95 $ x 128 semanas"
                datos_producto = re.sub(r'(\d+\s*\$)\s*Desde\s*x\s*(\d+\s*semanas)', r'\1 x \2', datos_producto)

                # Corregir el formato de "85 $ $ semanales Desde" a "85 $ semanales" con salto de línea
                datos_producto = re.sub(r'(\d+\s*\$)\s*\$\s*semanales\s*Desde', r'\1 semanales\n', datos_producto)

                # Corregir el formato de "110 $ semanales Desde" a "110 $ semanales" con salto de línea
                datos_producto = re.sub(r'(\d+\s*\$)\s*semanales\s*Desde', r'\1 semanales\n', datos_producto)

                if 'Incluye:' in datos_producto:
                    partes = datos_producto.split('Incluye:', 1)
                    primera_parte = partes[0].strip()
                    primera_parte = re.sub(r'^.*?(\d+\s*\$)', r'\1', primera_parte)
                    resto = partes[1].strip() if len(partes) > 1 else ''
                    datos_producto = f"Pago: {primera_parte}\nIncluye: {resto}"

                if 'Incluye:' not in datos_producto:
                    partes = datos_producto.split('\n', 1)
                    primera_parte = partes[0]
                    primera_parte = re.sub(r'^.*?(\d+\s*\$)', r'\1', primera_parte)
                    resto = partes[1] if len(partes) > 1 else ''
                    resto = resto.replace('Incluye:', '').strip()
                    datos_producto = f"Pago: {primera_parte}\nIncluye: {resto}"
                
                vigencia = vigencia.replace('al comprar solo con', '').strip()

            # Aplicar las transformaciones comunes independientemente del subtítulo
            datos_producto = datos_producto.replace('es la mejor opción para pagar menos', '')  # Eliminar esta frase
            datos_producto = datos_producto.replace('con tu', 'con tu préstamo Elektra')
            datos_producto = datos_producto.replace('Total a pagar con Préstamo Elektra', '\nTotal a pagar con Préstamo Elektra')

            # Eliminar frases no deseadas
            datos_producto = re.sub(r'Recuerda que el uso de casco.*', '', datos_producto)
            datos_producto = re.sub(r'Sku´s participantes:.*', '', datos_producto)
            datos_producto = re.sub(r'Sku´s\s+que no participan:.*', '', datos_producto)
            datos_producto = re.sub(r'Modelos seleccionados:.*', '', datos_producto)
            datos_producto = datos_producto.replace('Precio total a pagar a crédito:', '\nPrecio total a pagar a crédito:')
            
            # Verificar si el subtítulo es 'Producto'
            if subtitulo == 'Producto':
                categoria = titulos[0] if titulos else "Categoría Desconocida"
                subtitulo = f"Producto: Promoción {categoria}"
                content = f"Pago semanal: NA\nDescuento: {datos_producto}\nPago de contado: NA"
            else:
                # Si no es 'Producto', aplicar las transformaciones habituales
                match = re.search(r'^(.*?)(\d+ ?\$|\$\d+)', datos_producto)
                if match and titulos[0] != 'Planes':
                    subtitulo = f"Producto: {subtitulo} {match.group(1).strip()}"
                    datos_producto = datos_producto.replace(match.group(1), "").strip()
                    subtitulo = subtitulo.replace('$ ', '')
                elif titulos[0] == "Planes":
                    subtitulo = f"Producto: {subtitulo}"
                    datos_producto = datos_producto
                else:
                    subtitulo = f"Producto: {subtitulo}"
                
                # Revisar y ajustar las frases que preceden "abono semanal", "descuento" o "descuento en abono semanal"
                if 'descuento en abono semanal' in datos_producto:
                    datos_producto = datos_producto.replace('descuento en abono semanal', 'descuento en abono semanal')
                elif 'Descuento' in datos_producto or 'descuento' in datos_producto:
                    datos_producto = datos_producto.replace('descuento', 'descuento\nPago de contado:')
                if 'abono semanal' in datos_producto or 'Abono semanal' in datos_producto:
                    datos_producto = datos_producto.replace('abono semanal', 'abono semanal\nPago de contado:')
                
                if titulos[0] == "Cómputo":
                    datos_producto = re.sub(r"\s?L L É V A T E  B O C I N A D E  R E G A L O !\s?", "", datos_producto).strip()
                    datos_producto = datos_producto.replace("¡", "").replace("!", "").strip()
                    datos_producto = datos_producto.replace("¡ B O C I N A  D E  R E G A L O !", "").strip()
                    datos_producto = datos_producto.replace("B O C I N A  D E  R E G A L O", "").strip()
                    datos_producto = datos_producto.replace("+ T A B L E T  D E  R E G A L O ", "").strip()
                    subtitulo = subtitulo.replace("+ T A B L E T  D E  R E G A L O ", "").strip()
                    subtitulo = subtitulo.replace("L L É V A T E  B O C I N A", "").strip()

                # Asignar el contenido modificado
                content = f"{datos_producto}"
                if titulos[0] == "Equipos":
                    # Eliminar bonos de regalo transversales del campo "Pago semanal"
                    datos_producto = re.sub(r'Bono de regalo transversal.*?\.', '', datos_producto).strip()
                    datos_producto = re.sub(r'\$1,000\.', '', datos_producto).strip()

                    # Ajuste del campo "Pago semanal" para asegurar el formato correcto
                    datos_producto = re.sub(
                        r"(\$\d{1,3}(?:,\d{3})*)\s+(\$\d{1,3}(?:,\d{3})*)\s+\+\s+semanales",
                        r"\1 + \2 semanales",
                        datos_producto
                    )
                    datos_producto = re.sub(
                        r"\$(\d+)\s+\+\s+semanales te llevas un accesorio o un cargador\s+\$(\d{1,3}(?:,\d{3})*)",
                        r"$\2 + $\1 semanales te llevas un accesorio o un cargador",
                        datos_producto
                    )

                # Patrón para el formato de pago semanal
                datos_producto = datos_producto.replace("¡ B O C I N A  D E  R E G A L O !", "").strip()
                datos_producto = datos_producto.replace("+ T A B L E T  D E  R E G A L O ", "").strip()

                # Patrón para el formato de pago semanal
                pago_semanal_pattern = re.compile(r'(\$?\d+)\s*x\s*(\d+)\s*semanas\s*(\$\d{1,3}(?:,\d{3})*)\s*de\s*pago\s*inicial\s*(\d+)(.*)')

                # Aplicar el patrón y verificar coincidencia en el texto limpio
                match = pago_semanal_pattern.search(datos_producto)
                if match:
                    # Extraer los valores según el patrón
                    cantidad4 = match.group(4)  # Monto de pago semanal
                    semanas = match.group(2)     # Número de semanas
                    pago_inicial = match.group(3)  # Pago inicial
                    texto_adicional = match.group(5).strip()  # Cualquier texto adicional

                    # Verificar si hay un descuento en el texto adicional y formatear el mensaje
                    if 'Descuento' in texto_adicional or 'descuento' in texto_adicional:
                        texto_adicional = f"\nDescuento: {texto_adicional}"

                    # Formatear el texto final con el patrón deseado
                    nuevo_texto = f"${cantidad4} x {semanas} semanas {pago_inicial} de pago inicial"
                    datos_producto = datos_producto.replace(match.group(0), nuevo_texto + texto_adicional)
                    content = f"Pago semanal: {datos_producto}"

                # Revisar si 'enganche' aparece en datos_producto
                if 'enganche' in datos_producto:
                    datos_producto = datos_producto.replace('enganche', 'enganche\nDescuento:')
                    content = f"Pago semanal: {datos_producto}"

                # Revisar si hay 'pago inicial' en datos_producto
                elif re.search(r'pago inicial \d+', datos_producto):
                    datos_producto = re.sub(r'(pago inicial \d+)', r'\1\nDescuento:', datos_producto)
                    content = f"Pago semanal: {datos_producto}"
                
                subtitulo = re.sub(r"¡ B O C I N A  D E  R E G A L O !.*", "", subtitulo).strip()
                subtitulo = subtitulo.replace("S E G U N D A  P I E Z A H A S T A  - 7 0 %  E N  A B O N O %", "").strip()
                subtitulo = subtitulo.replace("S E G U N D A  P I E Z A H A S T A  - 6 0 %  E N  A B O N O %", "").strip()
                # Expresión regular para eliminar "¡ B O N O  D E  R E G A L O  D E  1 , 0 0 0 * !" y todo lo que sigue
                subtitulo = re.sub(r"¡\s*B\s*O\s*N\s*O\s*D\s*E\s*R\s*E\s*G\s*A\s*L\s*O\s*D\s*E\s*1\s*,\s*0\s*0\s*0\s*\*\s*!.*", "", subtitulo).strip()

                if titulos[0] == "Equipos":
                    subtitulo = subtitulo.replace("L L É V A T E  X I A O M I  R E D M I  A 3 D E  R E G A L O !", "").strip()
                    # Extraer características del producto como '* Cámara 12+12 Mpx Memoria: 128 GB'
                    caracteristicas_pattern = re.compile(r'\* Cámara\s+\d+\+\d+(?:\+\d+)?\s+Mpx\s+Memoria:\s+\d+\s+GB')
                    caracteristicas = caracteristicas_pattern.search(datos_producto)
                    
                    if caracteristicas:
                        caracteristicas_texto = caracteristicas.group(0)
                        # Eliminar las características del campo "Pago semanal"
                        datos_producto = datos_producto.replace(caracteristicas_texto, '').strip()
                        # Agregar las características al nombre del producto
                        subtitulo = f"{subtitulo} {caracteristicas_texto}"
                        subtitulo = subtitulo.replace("*", "").strip()
                        
                    content = f"Pago semanal: {datos_producto}"

            # Restante de las asignaciones
            url = urls[i] if i < len(urls) else f"{page_num}_Dummy{i}"
            categoria = f"Categoria: {re.sub(r'[^A-Za-zÁÉÍÓÚáéíóúÑñ\s]', '', titulos[0]).strip()}" if titulos else ""
            if categoria == "Categoria: ":
                categoria = ultima_categoria  # Usar la última categoría válida
            else:
                ultima_categoria = categoria

            categoria = categoria.replace("Paquete", "").strip()
            if categoria == 'Categoria: Planes':
                # Extrae el valor correcto para el pago semanal
                content = re.sub(r'\$\d+\s*(\d+\s*\$\s*semanales)', r'Pago: \1', content)

                # Asegúrate de que "Incluye:" esté en una nueva línea, en caso de estar seguido
                content = re.sub(r'(?<=semanales)\s*Incluye:', r'\nIncluye:', content)
            #cupon = cupones[i] if i < len(cupones) else "Sin Cupones"

            if sku:
                sku_num = "Sku: " + sku
                #url_line = f"Url: {url}"
                data = [sku_num, categoria, subtitulo, content, vigencia]
                nueva_data_productos.append([sku_num, categoria, subtitulo, content, vigencia])
                #guardar_informacion_a_discovery(titulos[0], f"{sku} {url}", data)
                guardar_informacion_a_elasticsearch(f"{sku}", data, bucket_name, carpeta_documentos_correcciones_bucket, carpeta_documentos_elastic_bucket)
    print(categorias_extraidas)
    generar_reporte_excel_general(bucket_name, carpeta_reportes_bucket)

def get_downloads_folder():
    # Obtener la carpeta de descargas según el sistema operativo
    if os.name == 'nt':  # Windows
        return str(Path.home() / 'Downloads')
    else:  # macOS y Linux
        return str(Path.home() / 'Downloads')

def limpiar_caracteres_especiales(texto):
    # Eliminar caracteres no imprimibles excepto saltos de línea, caracteres acentuados, $, %, /, comillas dobles, comillas dobles de cierre y el símbolo +
    return re.sub(r'[^a-zA-Z0-9áéíóúÁÉÍÓÚüÜñÑ.,;:!?()\[\]{}<>\-\n $%/""”+]+', '', texto)

def generar_reporte_excel_general(bucket_name, carpeta_reportes_bucket):
    # Extraer SKU y URL directamente desde `nueva_data_productos`
    skus = [data[0].replace("Sku: ", "") for data in nueva_data_productos]
    urls = [data[-1].replace("Url: ", "") for data in nueva_data_productos]  # Asume que el último campo es la URL

    # Crear la data en un DataFrame para todas las categorías en `nueva_data_productos`
    nueva_data_formateada = [limpiar_caracteres_especiales('\n'.join(data)) for data in nueva_data_productos]
    df = pd.DataFrame({
        'SKU': skus,  
        'URL': urls,  
        'Nueva Data Producto': nueva_data_formateada  
    })

    # Asegurarse de que tanto la columna 'SKU' como 'URL' sean de tipo string
    df['SKU'] = df['SKU'].astype(str).str.strip()
    df['URL'] = df['URL'].astype(str).str.strip()

    # Crear una columna para verificar si el SKU está contenido en la URL
    #df['URL Coincide con SKU'] = [sku in url for sku, url in zip(df['SKU'], df['URL'])]

    # Añadir una columna que indique si el SKU tiene una imagen asociada
    def tiene_imagen(sku):
        extensiones_imagen = ['jpeg']
        downloads_folder = get_downloads_folder()
        ruta_imagenes = os.path.join(downloads_folder, 'imagenes')
        for ext in extensiones_imagen:
            if os.path.exists(f"{ruta_imagenes}/{sku}.{ext}"):
                return True
        return False

    df['Tiene Imagen'] = df['SKU'].apply(tiene_imagen)

    # Definir la carpeta de descargas y el nombre del archivo consolidado
    downloads_folder = get_downloads_folder()
    nombre_archivo_excel = os.path.join(downloads_folder, "reporte_consolidado_categorias.xlsx")

    # Guardar el DataFrame como archivo Excel sin imágenes aún
    df.to_excel(nombre_archivo_excel, index=False)

    # Cargar el archivo Excel generado con openpyxl
    workbook = load_workbook(nombre_archivo_excel)
    sheet = workbook.active

    # Ajustar el ancho de las columnas "SKU", "URL Coincide con SKU" y "Tiene Imagen"
    col_sku = 1
    col_url_coincide = sheet.max_column - 1
    col_tiene_imagen = sheet.max_column 

    sheet.column_dimensions[sheet.cell(row=1, column=col_sku).column_letter].width = 30 
    sheet.column_dimensions[sheet.cell(row=1, column=col_url_coincide).column_letter].width = 30
    sheet.column_dimensions[sheet.cell(row=1, column=col_tiene_imagen).column_letter].width = 30

    # Ajustar el ancho de la columna de "Nueva Data Producto" para que sea 3 veces más ancho
    col_nueva_data = 3
    sheet.column_dimensions[sheet.cell(row=1, column=col_nueva_data).column_letter].width = 90 

    # Añadir imágenes a una nueva columna al final
    col_img = sheet.max_column + 1
    sheet.cell(row=1, column=col_img).value = "Imagen"

    # Añadir imágenes en la nueva columna para cada SKU en df
    for index, sku in enumerate(df['SKU'], start=2):
        imagen_path = None
        extensiones_imagen = ['jpeg']
        ruta_imagenes = os.path.join(downloads_folder, 'imagenes')

        for ext in extensiones_imagen:
            imagen_path = f"{ruta_imagenes}/{sku}.{ext}"
            if os.path.exists(imagen_path):
                break

        if imagen_path and os.path.exists(imagen_path):
            img = ExcelImage(imagen_path)
            img.width = 100
            img.height = 100
            fila = index
            sheet.row_dimensions[fila].height = img.height * 0.75 
            img_anchor = sheet.cell(row=fila, column=col_img).coordinate 
            sheet.add_image(img, img_anchor)

    # Aplicar formato condicional para resaltar en rojo las celdas con "False" en las columnas "URL Coincide con SKU" y "Tiene Imagen"
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # Aplicar formato condicional para resaltar en rojo las celdas con "False" en las columnas "URL Coincide con SKU" y "Tiene Imagen"
    # Comentado: Columna "URL Coincide con SKU"
    # for row in sheet.iter_rows(min_row=2, min_col=col_url_coincide, max_col=col_url_coincide, max_row=sheet.max_row):
    #     for cell in row:
    #         if cell.value == False:
    #             cell.fill = red_fill

    # Columna "Tiene Imagen"
    for row in sheet.iter_rows(min_row=2, min_col=col_tiene_imagen, max_col=col_tiene_imagen, max_row=sheet.max_row):
        for cell in row:
            if cell.value == False:
                cell.fill = red_fill

    # Marcar los SKUs duplicados en color naranja
    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

    # Identificar duplicados en la columna SKU
    sku_column = sheet['A'][1:]  # Column A, excluding header
    sku_values = [cell.value for cell in sku_column]
    duplicated_skus = {sku for sku in sku_values if sku_values.count(sku) > 1}

    # Aplicar el color naranja a los duplicados
    for cell in sku_column:
        if cell.value in duplicated_skus:
            cell.fill = orange_fill
    
    # Marcar en amarillo las celdas de "Nueva Data Producto" si el subtítulo comienza con "Producto: Promoción"
    # o si no cumplen con la estructura y no son de la categoría "equipos" o "planes"
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    estructura_requerida = ["Sku:", "Categoria:", "Producto:", "Pago semanal:", "Descuento:", "Pago de contado:", "Vigencia:"]
    
    def cumple_estructura(data):
        for campo in estructura_requerida:
            if not any(campo in item for item in data):
                return False
        return True
    
    # Define pattern for "Pago semanal"
    pago_semanal_pattern = re.compile(
        r'Pago semanal: \$?\d+ x \d+ semanas \$?\d+(,\d{3})* de pago inicial|Pago semanal: \d+ \$ x \d+ semanas con \d+% de enganche'
    )

    for index, data in enumerate(nueva_data_productos, start=2):  # Comienza en la fila 2
        subtitulo = data[2]  # Índice del subtítulo en `nueva_data_productos`
        categoria = data[1].replace("Categoria:", "").strip().lower()  # Remover etiqueta y espacios en blanco
        
        # Verificar si la categoría está vacía o si no cumple con la estructura
        if not categoria or subtitulo.startswith("Producto: Promoción") or (
            not cumple_estructura(data) and "equipos" not in categoria and "planes" not in categoria
        ) or (categoria != "planes" and not any(pago_semanal_pattern.search(line) for line in data)):
            cell = sheet.cell(row=index, column=3)  # Columna "Nueva Data Producto"
            cell.fill = yellow_fill

    # Guardar el archivo Excel con todas las validaciones e imágenes insertadas
    workbook.save(nombre_archivo_excel)
    print(f"Reporte consolidado guardado en {nombre_archivo_excel}")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    # Subir el archivo Excel al bucket de Google Cloud Storage
    with open(nombre_archivo_excel, "rb") as excel_file:
        excel_buffer = io.BytesIO(excel_file.read())
        excel_buffer.seek(0)  # Asegurarse de que el buffer esté al inicio

        # Subida al storage con el nombre que incluye fecha y hora
        #st.upload_text_buffer(bucket_name, carpeta_reportes_bucket, f"reporte_{timestamp}.xlsx", excel_buffer)
        print(f"El reporte de Excel '{nombre_archivo_excel}' ha sido subido exitosamente al bucket '{bucket_name}/{carpeta_reportes_bucket}'.")

